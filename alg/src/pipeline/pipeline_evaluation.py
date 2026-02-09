# -*- coding:utf-8 -*-
import os
import re
import sys
import json
import random
import argparse
import openpyxl
import concurrent.futures
from include.logger import log
from include.config import CommonConfig
from include.utils.llm_caller_utils import llm_call
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
es_user_config = CommonConfig['ES_USER']

rules_exp = {
    '意图理解一致性': """
关键点遗漏：未覆盖提问者问题的关键点，每处遗漏扣分1.0分；
限定条件不满足：未满足提问者问题的限定条件（如特定时间、地点、主体等），每个错误扣分1.0分；
回答不直接：仅间接回答提问者问题，而没有精确、直接回答，每个错误扣分1.0分；
回答不准确：考虑回答中的不准确之处，每处不准确扣1.0分；
冗余内容：生成结果中存在冗余的内容，即提问者未询问该内容，每个错误扣分0.5分（注意某些额外补充的内容可能是符合提问者意图的，仅在与冗余内容与作者意图不想符时扣分）
多角度分析：回答中提供了多角度的分析和观点，加1.0分。
""",
    '引证有效性': """
引证不一致：生成结果中的陈述和所引用的参考资料的描述不符（相同的事件时间、地点、人物、数字等有偏差），每个错误扣分1.0分；
幻觉：生成结果中存在不存在与参考资料的事实性信息，每个错误扣分1.0分
""",
    '结构条理性': """
结构不完整：考虑生成结果是否采取总分总/总分/分总结构。若仅直接堆砌信息，既没有总起也没有总结，则扣分1.0分；
文章结构清晰合理：若采取明确的总分总结构，且具备引语、背景等部分，额外加0.5分；
段落组织不当：每个段落应该围绕一个明确的主题展开。如果段落主题不明确或信息混乱，扣分1.0分；
层次不清晰：回答应按层次逐步展开，信息由浅入深。如果层次不清、信息堆砌扣分1.0分，情况严重者扣分2.0分；
详略不当：考虑文章中的重点内容与次要内容的篇幅分布，重点描述次要信息而对重要信息有所忽视的，扣分1.0分；
逻辑错误：考虑文章中的逻辑错误数量，每个错误扣分1.0分；
逻辑错误通常包括以下几类错误，但不限于：
矛盾：文章中存在逻辑矛盾，即相互冲突或不一致的陈述。
非因果关系：文章中建立了不正确的因果关系，导致不合理的结论。
逻辑跳跃：文章中出现了不明确的逻辑跳跃，读者难以理解或追踪作者的思维过程。
过渡自然：段落和句子之间过渡自然，回答流畅。如果过渡自然、连贯，额外加1.0分；
逻辑连接词：如果全文能够恰当使用连接词（如因此、而且、另外、同样、例如、然而），额外加分0.5分。
""",
    "内容相关性": """
内容偏离：考虑回答偏离提问主题的情况，每处偏离扣1.0分；
信息遗漏：考虑回答中遗漏的重要信息，每处遗漏扣1.0分；
过度概括：考虑回答过度概括而缺乏具体信息的情况，每处扣0.5分；
主题连贯性：考虑文章是否保持主题的连贯性，避免突变或无关内容的出现，如果有与主题无关的语句，视偏题程度扣1-2.0分；
"""
}


def decode_answer_from_llm_raw_output(llm_raw_result: str):
    """
    解析大模型输出的答案、思路以及其他可能的材料，不同的prompt输出格式不同、解析的内容也不同
    """
    pattern_answer = r"\[回答\]:*(.*?)(#+)?$"
    answer, thought = None, None
    answer = re.search(pattern_answer, llm_raw_result, re.DOTALL)
    return answer, thought


def generate_gpt_prompt(answer_path):
    """
        参数：
        answer_path : str
            检索生成的结果文件路径
        prompt_path : str
            调GPT-4的prompt文件路径

        prompt文件中每条数据结构：
        {
            "prompt": "prompt",
            "_id": "5249a7d0fba7478921e418eb370415d4adf10f611b34867ee0c04860662a1b21multi_hop_general_xinhua0_05131716303693.6842785",
            "dimension": "指令服从"
        }
    """
    prompts = []
    with open(answer_path,  'r', encoding='utf-8') as file:
        answers = json.load(file)
        for answer in answers:
            if 'output_llm' in answer["llm_sft_data"] and len(answer["llm_sft_data"]['output_llm'][0]) > 10:
                q = "【问题】: \n" + answer["content"]["question"].strip() + "\n"
                a, _ = decode_answer_from_llm_raw_output(answer["llm_sft_data"]["output_llm"][0].lstrip().rstrip())
                if a:
                    res = a.group(1)
                    # 定义需要被过滤掉的模式
                    pattern = r'\[[a-zA-Z0-9]+\]'
                    # 使用正则表达式进行替换
                    a = re.sub(pattern, '', res)
                else:
                    a = answer["llm_sft_data"]["output_llm"][0].lstrip().rstrip()
                a = re.sub(r'\[\w+\]', '', a).strip().lstrip().rstrip()
                a = "【回答】: \n" + a + "\n"

                r = answer["llm_sft_data"]['input_llm'].split("日\n\n参考资料：")[1]
                r = "【引证材料】: \n" + r.split('结构化模版：\n以下是一些使答案更加清晰')[0].strip()

                for key, value in rules_exp.items():
                    if key == '引证有效性':
                        prompt = "假设你是一位文章质量检查员，请帮我评估回答的{}，我将给你用户的问题、参考资料和最终的回答。满分是10分，具体打分规则如下：\n".format(key)
                    else:
                        prompt = "假设你是一位文章质量检查员，请帮我评估回答的{}，我将给你用户的问题和最终的回答。满分是10分，具体打分规则如下：\n".format(key)
                    prompt = prompt + value
                    prompt += "请严格遵循打分规则，并依照以上示例进行{}打分".format(key) + \
                                "答案以json格式给出：请分别给出打分的解释和你的打分。在打分解释中，避免任何潜在的偏见，并确保不会有除文本外其他因素影响你的判断。" \
                                "分数仅包含一个数字值，不要有额外的信息。\n\n\n" \
                                "示例输出: \n" + \
                                "{'打分解释': 'XXXXXXXX', '分数': 7}\n" + \
                                "请你根据以下所给的对话上下文，按照以上所给的评判标准和打分示例，对{}这个维度的答案质量进行打分：\n".format(key)
                    if key == '引证有效性':
                        prompt += q + a + r + "\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                "{'打分解释': 'XXXXXXXX', '分数': 7}\n, 你的打分结果: \n"
                    else:
                        prompt += q + a + "\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                              "{'打分解释': 'XXXXXXXX', '分数': 7}\n, 你的打分结果: \n"
                    prompts.append({
                        "prompt": prompt,
                        "_id": answer["_id"],  # 保存一下answer的id，方便以后其他的分析
                        "dimension": key,
                    })
            else:
                from include.logger import log
                log.warning("WRONG!!! answer: {}".format(answer))
    return prompts


def get_gpt_score(score_path, score_path_xlsx):
    """
        参数：
        score_path : str
            GPT-4打分结果文件路径
        score_path_xlsx : str
            统计打分结果保存Excel文件路径

        GPT-4打分结果文件只是每条数据比原prompt文件多了一个“gpt_ans”字段
    """
    task_scores = dict()
    for aspect in rules_exp.keys():
        task_scores[aspect] = []

    with open(score_path, 'r', encoding='utf-8') as file:
        res_lis = json.load(file)
    for obj in res_lis:
        dimension = obj.get("dimension")
        score = None
        if "gpt_ans" in obj:
            score = obj["gpt_ans"]['分数']
            try:
                score = int(score)
            except:
                try:
                    pattern = re.compile(r'\d+')
                    result = pattern.findall(score)
                    score = int(result[-1])
                except:
                    # print("分数格式有误！", score)
                    score = None
                    pass

        if score is not None and score in list(range(0, 11)):
            task_scores.get(dimension).append(score)

    wb = openpyxl.Workbook()
    sheet = wb.active

    # 写入标题行
    headers = list(task_scores.keys())
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)

    # 写入平均分数行
    for i, key in enumerate(task_scores.keys(), start=1):
        if len(task_scores[key]) != 0:
            average_score = sum(task_scores[key]) / len(task_scores[key])
        else:
            average_score = 0
        sheet.cell(row=2, column=i, value=average_score)

    # 保存 Excel 文件
    wb.save(score_path_xlsx)
    print("save in {}".format(score_path_xlsx))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="结果评估")
    parser.add_argument("--parallel_batch", default=18, help="")
    parser.add_argument("--seed", default=3, help="")
    parser.add_argument("--exp_name", default='AINews_answer_GraTAG_new_0830', help="")
    parser.add_argument("--eval_model", default='gpt-4-0125-preview', help="")
    parser.add_argument("--rule_name", default='rule-1', help="")

    # 解析命令行参数
    args = parser.parse_args()

    random.seed(args.seed)
    answers = generate_gpt_prompt(
        answer_path="../data/result/{}.json".format(args.exp_name)
    )

    prompt_path = "../data/prompt_result/prompt_{}_{}_{}.json".format(args.rule_name, args.eval_model, args.exp_name)
    eval_path = "../data/eval_result/score_{}_{}_{}.xlsx".format(args.rule_name, args.eval_model, args.exp_name)
    # 创建目录
    os.makedirs(os.path.dirname(prompt_path), exist_ok=True)
    os.makedirs(os.path.dirname(eval_path), exist_ok=True)


    def process_answer(answer, model_name, answer_index):
        while True:
            response_origin = llm_call(query=answer["prompt"], model_name=model_name, n=1, clog=log)
            if response_origin:
                try:
                    response = response_origin.replace("json", '').replace('```', '')
                    response = response.strip()
                    data = eval(response)
                    answer["gpt_ans"] = data
                    assert '分数' in data, "'分数' in data"
                    log.info("answer_index: {}, gpt_ans: {}".format(answer_index, response))
                    return answer
                except:
                    log.error("WRONG!!!! prompt: {}, {}".format(answer["prompt"], response_origin))


    # 并行处理答案
    answers = answers[:]
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.parallel_batch) as executor:
        future_to_answer = {executor.submit(process_answer, answer, args.eval_model, answer_index): answer for answer_index, answer in
                            enumerate(answers)}
        for future in concurrent.futures.as_completed(future_to_answer):
            answer = future_to_answer[future]
            try:
                answer = future.result()
            except Exception as exc:
                log.error(f'Answer {answer["prompt"]} generated an exception: {exc}')

    with open(prompt_path, 'w', encoding="utf-8") as f:
        json.dump(answers, f, indent=4, ensure_ascii=False)

    get_gpt_score(prompt_path, eval_path)