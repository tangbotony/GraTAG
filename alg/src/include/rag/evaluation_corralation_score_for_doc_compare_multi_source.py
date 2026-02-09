import json
import time
import copy
from tqdm import tqdm  # 进度条库
import logging
import numpy as np
import pandas as pd
import concurrent.futures
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from loguru import logger as log
from concurrent.futures import ThreadPoolExecutor
from include.utils.llm_caller_utils import llm_call


def queryDocCorrelation(query, doc):
    """
        判断query与材料是否有关联
        :param query: 查询query
        :param doc: 文档
        :return: True False bool
    """
    task_prompt_template = """你将被提供一个query和一段材料，判断材料信息是否与query有关，如果有关，则回答"true",否则返回"false"。
query: {}
材料：{}"""
    prompt = task_prompt_template.format(query, doc)
    ans = 0
    while True:
        try:
            response = llm_call(prompt, model_name='gpt-4o', clog=logging.log, n=1)
            if 'true' in response or 'True' in response:
                ans = 1
            break
        except Exception as e:
            print('err:', e)
            break
    return ans


def get_QA_score(res_i):
    def get_QA_score_doc_i(doc, query_i=res_i['test_query_i']):
        return {
            "relation_score": queryDocCorrelation(query_i, doc['description']),
            "origin_score": doc.get('rerank_score', None)
        }

    with ThreadPoolExecutor(50) as executor:
        ans_list = list(executor.map(get_QA_score_doc_i, res_i['sorted_sources'].values()))

    relation_score_list = [ans_i['relation_score'] for ans_i in ans_list]
    origin_score_list = [ans_i['origin_score'] for ans_i in ans_list]

    print("Query {} 的平均分是: {}".format(res_i['test_query_i'], np.mean(relation_score_list)))
    return {
        "query": res_i['test_query_i'],
        "category_c1": res_i.get('c1'),
        "category_c2": res_i.get('c2'),
        "avg_score": np.mean(relation_score_list),
        "scores": relation_score_list,
        "origin_scores": origin_score_list,
        "sources": res_i['sources']
    }


def get_sorted_sources(query_i_reference_source_j, query_i):
    from include.rag.rag_recall_agent import RagRecall
    from include.context.reference_controller import ReferenceController


    final_searched_items = []
    all_content_all_url = set()
    for item in query_i_reference_source_j:
        if item['content'] not in all_content_all_url:
            all_content_all_url.add(item['content'])
            final_searched_items.append(item)
    my_rag_recall = RagRecall(
        user_info={
            'application': 'GraTAG',
            'session_id': "test_rag_recall_session_id",
            'question_id': "test_rag_recall_question_id",
            'request_id': time.time()
        },
        query=query_i, logger=log,
        credible='false', auto_router_source='false',
        query_reinforce=False,
        similarity_config={
            "method": 'rerank_base',  # base/llm/rerank/rerank_base
            "is_filter": True,
            "simi_bar": 0.0,
            "model_name": 'gpt-4o',
            "top_n": 30
        },
        auto_retrieval_range=True,
        min_num_chunks=3
    )
    # chunk分片
    chunk_searched_items = my_rag_recall._split_chunks(final_searched_items)

    # 相似chunk过滤
    related_chunks = my_rag_recall._chunk_rank(chunk_searched_items)

    # 整理最终检索到的材料
    reference = ReferenceController()
    references_data_base = my_rag_recall._generate_final_database(related_chunks)
    if references_data_base:
        reference.add(references_data_base)
    log.debug("过滤后的chunk数量:{}".format(reference.get_num_references()))
    my_rag_recall._data_base = reference.get_all()
    sorted_sources = my_rag_recall.get_data_base()
    return sorted_sources['use_for_check_items']


if __name__ == '__main__':
    def generate(file_path):
        with open('{}.json'.format(file_path), 'r', encoding='utf-8') as file:
            data_file = json.load(file)

            def process_query(query_i, query_i_reference):
                results = []
                for source_name, query_i_reference_source_j in query_i_reference.items():
                    if isinstance(query_i_reference_source_j, list):
                        sorted_sources = get_sorted_sources(query_i_reference_source_j, query_i)
                        results.append({
                            'test_query_i': query_i,
                            'sorted_sources': sorted_sources,
                            'c1': query_i_reference['c1'],
                            'c2': query_i_reference['c2'],
                            'sources': source_name
                        })
                return results

            def process_data_in_parallel(data_file):
                res_lis = []
                with concurrent.futures.ThreadPoolExecutor(5) as executor:
                    # 创建任务
                    futures = [executor.submit(process_query, query_i, query_i_reference) for query_i, query_i_reference
                               in data_file.items()]
                    # 收集结果
                    for future in tqdm(concurrent.futures.as_completed(futures), total=len(futures),
                                       desc="Processing Queries"):
                        res_lis.extend(future.result())
                return res_lis

            # 调用并行处理的函数
            res_lis = process_data_in_parallel(data_file)

        with open("{}_sources.json".format(file_path), 'w', encoding="utf-8", errors="ignore") as f:
            json.dump(res_lis, f, indent=4, ensure_ascii=False)

    scores = []
    # generate("rag_evaluation/rag_result_source_routing_exp")

    with open('{}_sources.json'.format("rag_evaluation/rag_result_source_routing_exp"), 'r', encoding='utf-8') as file:
        data_sources = json.load(file)

    def evaluate(data_sources, file_path):
        def process_source(res_i):
            return get_QA_score(res_i)

        # 使用 ThreadPoolExecutor 进行并行处理
        with ThreadPoolExecutor(max_workers=10) as executor:
            # 使用 tqdm 包裹 executor.map，total 参数是总任务数量
            scores = list(tqdm(executor.map(process_source, data_sources), total=len(data_sources)))

        with open("{}_scores.json".format(file_path), 'w', encoding="utf-8") as f:
            json.dump(scores, f, indent=4, ensure_ascii=False)

    # evaluate(data_sources[:], "rag_evaluation/rag_result_source_routing_exp")

    with open('{}_scores.json'.format("rag_evaluation/rag_result_source_routing_exp"), 'r', encoding='utf-8') as file:
        data_scores = json.load(file)

    excel_file = '1023_comparison_results_between_sources.xlsx'
    new_source = list()
    for source_score_i in data_scores:
        source_score_i['平均材料相关度打分'] = source_score_i['avg_score']
        source_score_i['材料检索个数'] = len(source_score_i['scores'])
        new_source.append({
            "query": source_score_i['query'],
            "source": source_score_i['sources'],
            "category_c1": source_score_i['category_c1'],
            "category_c2": source_score_i['category_c2'],
            "平均材料相关度打分": source_score_i['平均材料相关度打分'],
            "材料检索个数": source_score_i['材料检索个数']
        })

    df = pd.DataFrame(new_source)
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Comparison')

    scores = data_scores
    score_dict = dict()
    for score_i in scores:
        if score_i['sources'] not in score_dict:
            score_dict[score_i['sources']] = list()
        score_dict[score_i['sources']].append(score_i)

    dataframes = []
    for source_name, source_score_i in score_dict.items():
        df = pd.DataFrame(source_score_i)
        df['平均材料相关度打分_{}'.format(source_name)] = df['avg_score']
        df['材料检索个数_{}'.format(source_name)] = df['scores'].apply(len)
        df = df.drop(columns=['avg_score', 'scores', 'origin_scores'])
        dataframes.append(df)

    # 使用 pd.concat 按列拼接 DataFrame
    merged_df = pd.concat(dataframes, axis=1)

    # 去除重复的 "query" 和 "检索词" 列，保留第一列出现的列
    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

    # 保存到Excel文件
    excel_file = '1023_comparison_results_between_sources.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, index=False, sheet_name='Comparison')

        # 加载Excel文件以应用样式
        workbook = writer.book
        worksheet = writer.sheets['Comparison']

        # 定义红色填充样式
        red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

        # 获取所有包含 "平均材料相关度打分" 和 "材料检索个数" 的列
        score_cols = [col for col in merged_df.columns if '平均材料相关度打分' in col]

        # 遍历每一行，为每个 query 的最高值标红
        for index, row in merged_df.iterrows():
            # 过滤掉非数值数据
            score_data = pd.to_numeric(row[score_cols], errors='coerce')

            max_score_col = score_data.idxmax()

            if pd.notna(max_score_col):
                worksheet.cell(row=index + 2, column=merged_df.columns.get_loc(max_score_col) + 1).fill = red_fill
    print(f"数据已成功保存到 {excel_file} 并完成标红。")







