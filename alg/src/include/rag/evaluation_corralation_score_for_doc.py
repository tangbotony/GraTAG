import json
import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor
from include.utils.llm_caller_utils import llm_call


def queryDocCorrelation(query, doc):
    """
        判断query与材料是否有关联
        :param query: 查询query
        :param doc: 文档
        :return: True False bool
    """
    task_prompt_template = """你将被提供一个query和一段材料，判断材料信息是否与query有关，如果有关，则回答"true",否则返回"false"。
query: {}
材料：{}"""
    prompt = task_prompt_template.format(query, doc)
    ans = 0
    while True:
        try:
            response = llm_call(prompt, model_name='gpt-4o', clog=logging.log, n=1)
            if 'true' in response or 'True' in response:
                ans = 1
            break
        except Exception as e:
            print('err:', e)
            break
    return ans


def get_QA_score(res_i):
    def get_QA_score_doc_i(doc, query_i=res_i['test_query_i']):
        return {
            "relation_score": queryDocCorrelation(query_i, doc['description']),
            "origin_score": doc.get('rerank_score', None)
        }

    with ThreadPoolExecutor(50) as executor:
        ans_list = list(executor.map(get_QA_score_doc_i, res_i['sorted_sources'].values()))

    relation_score_list = [ans_i['relation_score'] for ans_i in ans_list]
    origin_score_list = [ans_i['origin_score'] for ans_i in ans_list]

    print("Query {} 的平均分是: {}".format(res_i['test_query_i'], np.mean(relation_score_list)))
    return {
        "query": res_i['test_query_i'],
        "category": res_i.get('category'),
        "guidance_type": res_i.get('guidance_type', ''),
        "avg_score": np.mean(relation_score_list),
        "scores": relation_score_list,
        "origin_scores": origin_score_list,
        "time": res_i.get("time", 0.0)
    }


if __name__ == '__main__':
    files = [
        "0826_all_base_results_rerank_base-2",
        "0826_all_base_results_rerank_sft_0829_all_old_database-2",
        "0826_all_base_results_rerank_base_all_old_database-2"
    ]
    names = [
        "gpt-4o",
        "rerank_sft_0902_old_all",
        "rerank_base"
    ]

    def evaluate(file_path):
        with open('{}.json'.format(file_path), 'r', encoding='utf-8') as file:
            res_lis = json.load(file)
            res_lis = res_lis[:]
            for res in res_lis:
                if isinstance(res['sorted_sources'], list):
                    res['sorted_sources'] = res['sorted_sources'][:30]
            res_lis = res_lis[:]

        def process_source(res_i):
            return get_QA_score(res_i)

        # 使用 ThreadPoolExecutor 进行并行处理
        with ThreadPoolExecutor(1) as executor:
            scores = list(executor.map(process_source, res_lis))

        with open("score_{}.json".format(file_path), 'w', encoding="utf-8") as f:
            json.dump(scores, f, indent=4, ensure_ascii=False)

    # scores = []
    # for file_path in files[2:]:
    #     evaluate(file_path)

    scores = []
    for file_path in files:
        with open('score_{}.json'.format(file_path), 'r', encoding='utf-8') as file:
            res = json.load(file)
            scores.append(res)

    # 将数据转换为DataFrame
    dataframes = []
    for score_i, file_path in zip(scores, names):
        df = pd.DataFrame(score_i)
        df['平均材料相关度打分_{}'.format(file_path)] = df['avg_score']
        df['材料检索个数_{}'.format(file_path)] = df['scores'].apply(len)
        df = df.drop(columns=['avg_score', 'scores', 'time'])
        dataframes.append(df)

    # 使用 pd.concat 按列拼接 DataFrame
    merged_df = pd.concat(dataframes, axis=1)

    # 去除重复的 "query" 和 "检索词" 列，保留第一列出现的列
    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

    # 将数据转换为DataFrame
    dataframes = []
    for score_i, file_path in zip(scores, names):
        df = pd.DataFrame(score_i)
        df['平均材料相关度打分_{}'.format(file_path)] = pd.to_numeric(df['avg_score'], errors='coerce')
        df['材料检索个数_{}'.format(file_path)] = df['scores'].apply(len)
        df = df.drop(columns=['avg_score', 'scores', 'time'])
        dataframes.append(df)

    # 保存到Excel文件
    excel_file = '0927_new_comparison_results_old_data_time-2.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, index=False, sheet_name='Comparison')

        # 加载Excel文件以应用样式
        workbook = writer.book
        worksheet = writer.sheets['Comparison']

        # 定义红色填充样式
        red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

        # 获取所有包含 "平均材料相关度打分" 和 "材料检索个数" 的列
        score_cols = [col for col in merged_df.columns if '平均材料相关度打分' in col]

        # 遍历每一行，为每个 query 的最高值标红
        for index, row in merged_df.iterrows():
            # 过滤掉非数值数据
            score_data = pd.to_numeric(row[score_cols], errors='coerce')

            max_score_col = score_data.idxmax()

            if pd.notna(max_score_col):
                worksheet.cell(row=index + 2, column=merged_df.columns.get_loc(max_score_col) + 1).fill = red_fill
    print(f"数据已成功保存到 {excel_file} 并完成标红。")







