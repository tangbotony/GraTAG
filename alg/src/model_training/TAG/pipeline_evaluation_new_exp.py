# -*- coding:utf-8 -*-
import os
import re
import sys
import json
import random
import argparse
import openpyxl
import concurrent.futures
# from include.logger import log
# from include.utils.llm_caller_utils import llm_call
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
import logging
# from chat_with_GPT import chat_with_gpt

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
log = logging.getLogger(__name__)

key = "xx"
model = "gpt-4o"

def chat_with_gpt(query: str,
                  key: str = key,
                  model: str = model,
                  system_message: str = None,
                  temperature: float = 0.2,
                  retry_time: int = 5,
                  json_mode: bool = False
                  ):
    url = "xx"
    if system_message:
        message = [
            {
                "role": "system",
                "content": system_message
            },
            {
                "role": "user",
                "content": query
            }
        ]
    else:
        message = [
            {
                "role": "user",
                "content": query
            }
        ]
    payload = {
        "model": model,
        "messages": message,
        "temperature": temperature
    }
    if json_mode:
        payload.update(response_format={"type": "json_object"})
    payload = json.dumps(payload)
    headers = {
        'Authorization': 'Bearer {}'.format(key),
        'Content-Type': 'application/json',
    }
    count = 0
    while True:
        try:
            response = requests.request("POST", url, headers=headers, data=payload, timeout=300)
            result = json.loads(response.text)["choices"][0]["message"]["content"]
            break
        except Exception as e:
            count = count + 1
            print(e)
            if count > retry_time:
                raise Exception('ReturnCode.LLM_ERROR')
    return result

rules = {
        "指令服从": "我要你担任评分专家，从指令服从的角度对大模型的回答进行评分，满分为十分，打分标准如下：\n"
                    "10分：答即所问，答案是对用户提示词的直接回应，用户提示词中的所有要求或问题都能从答案中找到一一对应的信息，无论信息正确与否。若引证素材库中不存在与用户提示词对应的信息，答案则类似“我不知道”，而不是牵强附会。\n"
                    "5分：用户提示词中的部分要求或问题在答案中找不到针对性的信息，或者回答方式与问句类型不匹配，如对于是非问句没有回答“是”或“不是”、对于选择问句没有挑选问句中的某一个选项作答，而是需要用户从答案中进行推理才能得出所需信息。或者，有偷换概念的情况，如答案中把用户提示词里的关键名词性短语进行了近义替换。\n"
                    "0分：答非所问，用户提示词中的每一项要求或问题都从答案中找不到明确对应信息。或者，答案忽略了用户提示词中的限定性条件，如用户提示词设定有时间、空间范围，但答案混有限定范围以外的信息。或者，即便引证素材库中不存在与用户提示词对应的信息，仍作出似是而非的回答。\n",
                    # "以下是我给你的打分示例，请你参考："
                    # "问题：三星堆博物馆与原神的关系。"
                    # "打分示例：4\n\n指令服从存在较大偏差，着重回答了三星堆博物馆的特点与三星堆文化的价值，这与原问题想要了解两个主体间的关系的意图不符合。除了《原神》与三星堆⽂化的关联部分较为符合意图，其他三个部分提供的信息不符合提问者的需求，如三星堆博物馆的创新展陈与⽂化传播、三星堆⽂化的独特性与中华⽂明的关系、三星堆⽂化的独特性与中华⽂明的关系。\n",
        "引证溯源": "我要你担任评分专家，从引证溯源的角度对大模型的回答进行评分，满分为十分，打分标准如下：\n"
                    "10分：返回的答案中，每个句子都能从答案所附的“参考资料”中找到对应的信息源头。同时，“参考资料”列出的每一条信息源头都没有偏离用户提示词的核心要义。\n"
                    "8分：返回的答案中，75%以上的句子都能从答案所附的“参考资料”中找到对应的信息源头。同时，“参考资料”列出的信息源头中，仅有不到25%与用户提示词的核心要义无关。\n"
                    "6分：返回的答案中，有50%至75%的句子都能从答案所附的“参考资料”中找到对应的信息源头。同时，“参考资料”列出的信息源头中，有50%至75%紧扣用户提示词的核心要义。\n"
                    "4分：返回的答案中，仅10%至50%的句子能从答案所附的“参考资料”中找到对应的信息源头，或者，“参考资料”中50%至90%的信息源头跟用户提示词的核心要义关联不大。\n"
                    "0分：答案所附的“参考资料”内容为空值，或者“参考资料”列出的每一条信息源头跟用户提示词的核心要义之间的关联度都极低。\n",
                    # "以下是我给你的打分示例，请你参考："
                    # "问题：请梳理近两年来关于校园防欺凌的相关报道，提炼报道重点"
                    # "打分示例：4\n\n没有列举任何参考资料，其中“教育部统计显示，仅今年5⽉⾄8⽉就上报了68起校园欺凌事件”没有明晰的出处，经查证后发现数据来源为2016年，非常欠缺时效性。",
        "准确与否": "我要你担任评分专家，从准确与否的角度对大模型的回答进行评分，满分为十分，打分标准如下：\n"
                    "10分：对于用户提示词涉及的人、事、物，答案完全符合基本事实及逻辑关联，准确填补了用户提示词里的未知信息，无论答案是否精炼以及是否包含超越用户提示词以外的延展信息。\n"
                    "8分：答案中至少75%的信息符合基本事实及逻辑关联。\n"
                    "6分：答案中50%至75%的信息符合基本事实及逻辑关联。\n"
                    "4分：答案中至多只有50%的信息符合基本事实及逻辑关联。\n"
                    "0分：答案完全偏离基本事实，或者逻辑关联完全错乱。\n",
                    # "以下是我给你的打分示例，请你参考："
                    # "问题：总书记引⽤过哪些苏轼的作品"
                    # "打分示例：4\n\n回答这个问题比较合理的逻辑是按照引用的时间顺序，或是按照引用内容的类别等有条理地展开，但回答采用了混乱的逻辑，习近平总书记在三苏祠的引⽤、治国理政中的引⽤、⽂化⾃信和传统⽂化的引⽤以及其他引⽤实例四个部分，不满足“互斥且穷举“的要求，而且存在明显的不并列问题，三苏祠这部分是引用的时间/地点分类，但后两个部分则是按照内容分类。此外，详略不够得当，如仅把一次引用单独作为小标题，却忽略了其他场合，侧重点不对。\n",
        "信息延展": "我要你担任评分专家，从信息延展的角度对大模型的回答进行评分，满分为五分，打分标准如下：\n"
                    "5分：答案聚焦于用户提示词的要求或问题，仅含有不到25%的延展信息，且延展信息具有补充说明性质或较高启发性。\n"
                    "3分：针对用户提示词的回应信息散乱分布在答案中，答案有25%至75%的信息都不是在直接回应用户提示词中的要求或问题。\n"
                    "1分：答案围绕用户提示词的要求或问题东拉西扯，其中延展信息在整体答案中的占比超过75%。\n",
                    # "以下是我给你的打分示例，请你参考："
                    # "问题：上海⻢拉松路线经过苏州河么。"
                    # "打分示例：6\n\n整体上回答了问题，如“上海⻢拉松路线经过苏州河”，并配合具体例子来说明结论。但存在一些不必要的次要内容：此外，苏州河沿岸的滨⽔空间也被⽤于各种⽂体活动，成为市⺠休闲和运动的好去处。苏州河的治理和环境改善使其成为适合举办各类体育赛事的场所。近年来，苏州河的⽔质已经达到举办赛艇⽐赛的要求，并且在苏州河上举办的赛艇公开赛也充分展示了城市建设的成果。苏州河两岸的绿⾊⽣态空间和配套服务⽔平的提升，使其成为新的⽂化活动集聚区，进⼀步增强了城市的⼈⽂⽓息。这部分信息与问题不相关，对于用户的价值非常低。\n"
}

rules_strict = {
    '意图理解一致性': """
关键点遗漏：未覆盖提问者问题的关键点，每处遗漏扣分1.0分；
限定条件不满足：未满足提问者问题的限定条件（如特定时间、地点、主体等），每个错误扣分1.0分；
回答不直接：仅间接回答提问者问题，而没有精确、直接回答，每个错误扣分1.0分；
回答不准确：考虑回答中的不准确之处，每处不准确扣1.0分；
冗余内容：生成结果中存在冗余的内容，即提问者未询问该内容，每个错误扣分0.5分（注意某些额外补充的内容可能是符合提问者意图的，仅在与冗余内容与作者意图不想符时扣分）
多角度分析：回答中提供了多角度的分析和观点，加1.0分。
""",
    '引证有效性': """
引证不一致：生成结果中的陈述和所引用的参考资料的描述不符（相同的事件时间、地点、人物、数字等有偏差），每个错误扣分1.0分；
幻觉：生成结果中存在不存在与参考资料的事实性信息，每个错误扣分1.0分
""",
    '结构条理性': """
结构不完整：考虑生成结果是否采取总分总/总分/分总结构。若仅直接堆砌信息，既没有总起也没有总结，则扣分1.0分；
文章结构清晰合理：若采取明确的总分总结构，且具备引语、背景等部分，额外加0.5分；
段落组织不当：每个段落应该围绕一个明确的主题展开。如果段落主题不明确或信息混乱，扣分1.0分；
层次不清晰：回答应按层次逐步展开，信息由浅入深。如果层次不清、信息堆砌扣分1.0分，情况严重者扣分2.0分；
详略不当：考虑文章中的重点内容与次要内容的篇幅分布，重点描述次要信息而对重要信息有所忽视的，扣分1.0分；
逻辑错误：考虑文章中的逻辑错误数量，每个错误扣分1.0分；
逻辑错误通常包括以下几类错误，但不限于：
矛盾：文章中存在逻辑矛盾，即相互冲突或不一致的陈述。
非因果关系：文章中建立了不正确的因果关系，导致不合理的结论。
逻辑跳跃：文章中出现了不明确的逻辑跳跃，读者难以理解或追踪作者的思维过程。
过渡自然：段落和句子之间过渡自然，回答流畅。如果过渡自然、连贯，额外加1.0分；
逻辑连接词：如果全文能够恰当使用连接词（如因此、而且、另外、同样、例如、然而），额外加分0.5分。
""",
    "内容相关性": """
内容偏离：考虑回答偏离提问主题的情况，每处偏离扣1.0分；
信息遗漏：考虑回答中遗漏的重要信息，每处遗漏扣1.0分；
过度概括：考虑回答过度概括而缺乏具体信息的情况，每处扣0.5分；
主题连贯性：考虑文章是否保持主题的连贯性，避免突变或无关内容的出现，如果有与主题无关的语句，视偏题程度扣1-2.0分；
"""
}


# rules_8 = {
# "引证有效性": """假设你是一位阅卷老师，给定一个问题，请帮我评估给定【引证有效性】。满分是10分，具体打分规则如下：
# 1. 文章中的陈述和所给出的参考资料的描述不符（相同的事件时间、地点、人物、数字等有偏差），每个错误扣分1.0分；
# 2. 通识事实错误：明显虚构的人物、地点、事件，比如不存在的国家领导人、错误的地区名、错误的所属关系、违背常识的语句等，每个错误扣分1.0分；
#
# 举个例子，对于下面的问答, 我们会在文章后附上可能的参考资料，评价引证有效性的示例如下：
#
# 【问题】：“介绍2023年诺贝尔奖得奖情况”
# 【回答】：“2023诺贝尔奖新鲜出炉，总体得奖情况如下：首先，来自法国的科学家L’Huillier因为他们在阿秒脉冲光方面所做出的贡献被授予了2023诺贝尔奖。这一实验方法被用于产生阿秒光脉冲，以研究物质中电子运动学。
#
# 此外，奖项也授予了很多杰出的作家。坦桑尼亚作家阿卜杜勒-拉扎克·古尔纳以其作品在揭示殖民主义的影响和关注难民命运方面的卓越贡献，获得了诺贝尔文学奖。他的作品《天堂》《启程的记忆》《来生》等都备受推崇。奥地利作家彼得·汉德克也因其语言独创性与影响力的作品获得了诺贝尔文学奖。
#
# 在经济学领域，Imbens因其在劳动经济学及实证方法研究领域的突出贡献，被授予了2023年的诺贝尔经济学奖。
#
# 【引证材料】：资料0: L’Huillier,以表彰他们在阿秒脉冲光方面所做出的贡献。
#
# 资料1: L’Huillier),以表彰他们“用于产生阿秒光脉冲以研究物质中电子动力学的实验方法”。他们将平分1100万瑞典克朗(约合730万元人民币)奖金。(总台记者
#
# 资料2: 坦桑尼亚作家阿卜杜勒-拉扎克·古尔纳。获奖理由:“毫不妥协但却富有同情心地洞穿了殖民主义的影响,同时关注被夹在不同文化和地缘鸿沟间难民的命运”。代表作品有《天堂》《启程的记忆》《来生》。
#
# 资料3: 2021年诺贝尔经济学奖授予来自美国的经济学家戴维·卡德（David Card）、乔舒亚·D·安格里斯特（Joshua D. Angrist）和吉多·W·因本斯（Guido W. Imbens），以表彰他们在劳动经济学及实证方法研究领域的突出贡献。
#
# ”
# 可以进行如下点评：
# 存在的问题：
# 1. 在资料1中写明产生阿秒脉冲用于研究物质中电子动力学，而文章中却说用阿秒研究物质中电子运动学，二者不一致，扣1.0分
# 2. 在资料3中，Imbens是2021的诺贝尔经济学奖得主，但是文章却将其写为2023年得主，时间不一致，扣1.0分
# 分数: 8.0
# 计算过程: 10-1.0-1.0 = 8.0
#
# 根据以下的问答内容和提供的引证资料，请问你会给以下文章的“引证有效性”打多少分，请严格遵循打分规则，并依照以上示例进行“引证有效性”打分，将最终分数表达为python可以解析的Json格式，一个字段是"存在的问题"，将问题以列表的形式表达，一个字段是"分数"，另一个字段是"计算过程"，不要给出Json之外的任何语句。文章如下：
# """,
#
#     "回答问题相关性": """假设你是一位阅卷老师，给定一个问题，请帮我评估给定【回答与所提问题的相关性】。满分是10分，具体打分规则如下：
# 紧扣问题：回答是否紧扣所提问题，如果偏离主题，每次偏离扣1.0分；
# 回答完整：回答是否全面覆盖问题的各个方面，如果回答不完整，视缺失程度扣1.0-3.0分；
# 偏题：回答是否包含与问题无关的信息，每次偏题扣1.0分；
#
# 举个例子，对于下面的提问和回答：
# 提问：“中国在人形机器人的研发、制造、应用等方面取得了哪些进展？存在哪些优势和短板？面临哪些壁垒和瓶颈？”
# 回答：“中国在人形机器人的研发方面取得了一定的进展，特别是在机器人的灵活性和智能算法上。例如，某些国产机器人已经能够完成一些基本的家务活动。然而，这并不是中国科技领域的全部。实际上，中国在量子计算领域也取得了突破，虽然这与问题的主题不完全相关。在人形机器人的应用方面，中国的市场潜力巨大，但同时也面临着技术成熟度不足和成本较高的问题。至于短板，我认为中国在高端制造领域还有待提高，这不仅体现在机器人技术上，也体现在其他高科技产品上。至于壁垒和瓶颈，我认为主要是国际市场的竞争激烈和技术保护主义。”
#
# 可以这样点评：
# 存在的问题：
# 1.回答部分紧扣人形机器人的研发和应用，但提到了与主题不相关的量子计算领域，扣1.0分。
# 2.回答没有具体的进展和面临的壁垒，也没有提及制造方面的信息，扣2.0分。
# 分数: 7.0
# 计算过程: 10-1.0-2.0 = 7.0
#
# 请严格遵循打分规则，并依照以上示例进行打分，将最终分数表达为python可以解析的Json格式，一个字段是"存在的问题"，将问题以列表的形式表达，一个字段是"分数"，另一个字段是"计算过程"，不要给出Json之外的任何语句。待打分的问答如下: \n""",
#
#
# "信息完整性": """{假设你是一位阅卷老师，给定一个问题，请帮我评估学生的回答结果的【信息完整性】。满分是10分，具体打分规则如下：
# 1. 回答没有全面涵盖用户询问的所有方面、给出详尽的信息。每少一个点扣1.0分；
# 2. 如对于询问具体次数的问题，不给出泛泛的“若干”“多次”这样的回答，要准确、具体。每处模棱两可的回答扣1.0分。
# 3. 回答绕弯子、不给出事实性的信息或例子，仅仅在泛泛而谈，扣2.0分；
#
# 举个例子，对于下面的问答：
# 【问题】：“中国为发展职业教育出台的重要举措有哪些？”
# 【回答】：“中国为发展职业教育出台了一系列重要举措，主要包括以下几个方面：
#
# 深化产教融合：政府强调产教融合对于提高职业教育质量、促进经济转型发展的重要性，并从优化职业教育供给结构、构建多元办学格局、协同推进产教深度融合等方面提出具体举措。
#
# 完善职业教育体系：中国计划到2025年基本建成现代职业教育体系，并推进技能型社会建设。到2035年，职业教育整体水平将进入世界前列，技能型社会基本建成。
#
# 强化职业教育类型特色：政府提出巩固职业教育类型定位，因地制宜推进职业教育与普通教育协调发展，并加快建立“职教高考”制度。
#
# 推进不同层次职业教育纵向贯通：提升中等职业教育办学质量，优化布局结构，实施中等职业学校办学条件达标工程，同时稳步发展职业本科教育。”
#
# 可以这样点评：
# 存在的问题：
# 1.回答中没有列出具体的政策文件或举措名称，使得信息缺乏一定的准确性和引用性。按照规则第3条，扣2.0分。
# 2.回答没有提供具体实施的时间节点或者具体措施的数量，虽然列举了举措的方向，但缺少了具体和详尽的信息。依据规则第1条，可以认为需要扣1.0分。
# 分数: 7.0
# 计算过程: 10-2.0-1.0 = 7.0
#
# 根据以下的问答内容，请问你会给以下问答的“信息完整性”打多少分，请严格遵循打分规则，并依照以上示例进行“信息完整性”打分，将最终分数表达为python可以解析的Json格式，一个字段是'存在的问题'，将问题以列表的形式表达，一个字段是'分数'，另一个字段是'计算过程'，不要给出Json之外的任何语句。待打分的问答如下："
# """,
#
#
#
#     "文本连贯性": """{假设你是一位阅卷老师，给定一个问题，请帮我评估学生的回答结果的【文本连贯性】。满分是10分，具体打分规则如下：
# 1. 如果段落之间的过渡不自然或生硬，每处扣1.0分；
# 2. 如果文章内的句子逻辑不清晰或前后矛盾，每处扣1.0分；
# 3. 如果文章结构不合理，导致内容不连贯，每处扣1.0分
# 4. 表述不清：考虑文章中的表述不清的地方，每处扣分1.0分；
# 5. 啰嗦和重复：考虑文章中啰嗦和重复的地方，每处扣分1.0分，注意某些重复的表达是必要的，仅在非必要重复时扣分；
#
# 举个例子，对于下面的问答：
# 问题：“分析狗咬女童的各方观点和最新进展”
# 回答：“此事发生后，不少观点认为犬主人应当负责。据常莎分析，如果犬主人未尽到看管义务，导致禁养犬或者烈性犬伤人，可能构成以危险方法危害公共安全罪、故意伤害罪或过失致人重伤罪，需承担相应的刑事责任。此外，由于本案中女童涵涵遭受的全身多处咬伤及右肾挫裂伤，涵涵需要面临巨额医疗费用，这也让众多人呼吁犬主人将这笔医疗费用承担下来。而广泛关注的是，本案的犬主人在明知饲养的是危险的烈性犬的情况下，却未履行看管义务，肆意放任烈性犬在公共场所外出，从而导致了这次烈性犬咬伤女童的事件，可能涉嫌过失致人重伤罪，或者过失以危险方法危害公共安全罪。值得警惕的是，狗主人可能最高可获刑七年。家属对此也表示了他们的担忧，希望能得到社会的帮助。一些公众人士表示，除了法律责任，犬主人更要承担受害人的损失。部分网友还认为，在找到涉事狗主人的情况下，狗主人和物业应当承担赔偿责任。”
#
# 可以这样点评：
# 存在的问题：
# 1.没有交代常莎是谁,属于表述不清，扣1.0分
# 2.文章两次提到'以危险方法危害公共安全罪'比较啰嗦,扣1.0分
# 3.'家属对此也表示了他们的担忧'未表达清楚是谁的家属,表述不清,扣1.0分
# 分数: 7.0
# 计算过程: 10-1.0-1.0-1.0 = 7.0
#
# 根据以下的问答内容，请问你会给以下问答的“文本连贯性”打多少分，请严格遵循打分规则，并依照以上示例进行“文本连贯性”打分，将最终分数表达为python可以解析的Json格式，一个字段是'存在的问题'，将问题以列表的形式表达，一个字段是'分数'，另一个字段是'计算过程'，不要给出Json之外的任何语句。待打分的问答如下："
# """
# }


rules_8 = {
    "回答问题相关性": "回答要直接针对用户的问题，避免无关内容，没有废话，不绕弯子。每处不相关的语句扣一分。",

    "指令完成程度": "如对于询问具体次数的问题，不给出泛泛的“若干”“多次”这样的回答，要准确、具体。每处模棱两可的回答扣一分。",

    "准确性": "如过用户在问题中限制了时段、地点、人物、事件，答案要满足问题中的要求。每处答非所问的语句扣一分。",

    "无幻觉": "回答的信息需要准确无误，特别是对事实类问题的回答。避免数字、数据事实性错误。每处数字、数据事实性错误扣一分。",

    "时效性": "对于追踪持续发展的新闻事件和紧急报道，给的信息要跟踪事件的最新报道。注意今天的日期是2024年11月3日，不具有时效性的问题，这一项不扣分；对于时效性问题，请判断回答的结果的时效性，时效性越差，酌情扣分越多。",

    "信息完整性": "回答应全面涵盖用户询问的所有方面、给出详尽的信息。回答要做到不需要额外再去检索就能了解事件的全貌，每缺失一处回答的必要点扣一分。",

    "表达清晰度": "表述需要清晰易懂，展现形式易读有条理，能够很快获得想要的信息。（表达不清晰的例子：明显应该用时间发展顺序的描述的却用了总分、并列等展现方式），每处表达不清晰的语句扣一分。",

    "文本连贯性": "回答要逻辑连贯，语句之间过渡流畅，遣词用句得当。每处表达不连贯的地方扣一分。",

    "创新性/独特性": "回答能提供独到的见解、有深度。该项基础分为6分，每个有创新性的观点或表达酌情加0.5-1分。"
}


rules_exp = rules_8


def decode_answer_from_llm_raw_output(llm_raw_result: str):
    pattern_answer = r"\[回答\]:*(.*?)(#+)?$"
    answer, thought = None, None
    answer = re.search(pattern_answer, llm_raw_result, re.DOTALL)
    return answer, thought


def generate_gpt_prompt(answer_path, rule_name):
    prompts = []
    with open(answer_path,  'r', encoding='utf-8') as file:
        answers = json.load(file)
        for answer in answers:
            if answer['output'] is not None and len(answer['output']) > 10:
                q = "【问题】: \n" + answer["question"].strip() + "\n"
                a, _ = decode_answer_from_llm_raw_output(answer["output"].lstrip().rstrip())
                if a:
                    res = a.group(1)
                    pattern = r'\[[a-zA-Z0-9]+\]'
                    a = re.sub(pattern, '', res)
                else:
                    a = answer["output"].lstrip().rstrip()

                a = re.sub(r'\[\w+\]', '', a).strip().lstrip().rstrip()
                a = "【回答】: \n" + a + "\n"
                r = "【引证材料】: \n" + answer['instruction'].split("日\n\n参考资料：")[1].split("结构化模版：")[0].strip()

                if rule_name == 'v0-base':
                    for key, value in rules.items():
                        prompt = "我现在有一项任务：大模型根据用户输入的原始问题，理解其真实意图，检索相关的内容生成合理的回答，并进行引证。\n"
                        prompt = prompt + value
                        prompt += "你需要利用自己的经验，思考该回答如何评分最能符合标准和描述。" \
                                    "答案以json格式给出：请分别给出打分的解释和你的打分。在打分解释中，避免任何潜在的偏见，并确保不会有除文本外其他因素影响你的判断。" \
                                    "分数仅包含一个数字值，不要有额外的信息。\n" \
                                    "示例输出: \n" \
                                    "{'打分解释': 'XXXXXXXX', '分数': 7}\n" \
                                    "请你根据以下所给的对话上下文，按照以上所给的评判标准和打分示例，对'引证材料'和'回答'质量进行打分：\n"
                        prompt += q + a + r + "示例输出: \n" \
                                    "{'打分解释': 'XXXXXXXX', '分数': 7}\n, 你的打分结果: \n"
                        prompts.append({
                            "prompt": prompt,
                            "dimension": key
                        })
                elif rule_name == 'rules_8':
                    for key, value in rules_exp.items():
                        if key == '引证有效性':
                            prompt = ''
                        else:
                            prompt = ''
                        prompt = prompt + value
                        if key == '引证有效性':
                            prompt += q + a + r + "\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                    "{'存在的问题': 'XXXXXXXX', '分数': 7, '计算过程': '10-1.0-1.0-1.0 = 7.0'}\n, 你的打分结果: \n"
                        else:
                            prompt += q + a + r +"\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                                  "{'存在的问题': 'XXXXXXXX', '分数': 7, '计算过程': '10-1.0-1.0-1.0 = 7.0'}\n, 你的打分结果: \n"
                        prompts.append({
                            "prompt": prompt,
                            "dimension": key,
                        })
                else:
                    for key, value in rules_exp.items():
                        if key == '引证有效性':
                            prompt = "假设你是一位文章质量检查员，请帮我评估回答的{}，我将给你用户的问题、参考资料和最终的回答。满分是10分，具体打分规则如下：\n".format(key)
                        else:
                            prompt = "假设你是一位文章质量检查员，请帮我评估回答的{}，我将给你用户的问题和最终的回答。满分是10分，具体打分规则如下：\n".format(key)
                        prompt = prompt + value
                        prompt += "请严格遵循打分规则，并依照以上示例进行{}打分".format(key) + \
                                    "答案以json格式给出：请分别给出打分的解释和你的打分。在打分解释中，避免任何潜在的偏见，并确保不会有除文本外其他因素影响你的判断。" \
                                    "分数仅包含一个数字值，不要有额外的信息。\n\n\n" \
                                    "示例输出: \n" + \
                                    "{'打分解释': 'XXXXXXXX', '分数': 7}\n" + \
                                    "请你根据以下所给的对话上下文，按照以上所给的评判标准和打分示例，对{}这个维度的答案质量进行打分：\n".format(key)
                        if key == '引证有效性':
                            prompt += q + a + r + "\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                    "{'打分解释': 'XXXXXXXX', '分数': 7}\n, 你的打分结果: \n"
                        else:
                            prompt += q + a + "\n\n\n请严格遵循打分规则，并依照以上示例进行“{}”打分。示例输出: \n" \
                                                  "{'打分解释': 'XXXXXXXX', '分数': 7}\n, 你的打分结果: \n"
                        prompts.append({
                            "prompt": prompt,
                            "dimension": key,
                        })
            else:
                # from include.logger import log
                log.warning("WRONG!!! answer: {}".format(answer))
    return prompts


def get_gpt_score(score_path, score_path_xlsx, rule_name):
    task_scores = dict()
    for aspect in rules_exp.keys():
        task_scores[aspect] = []

    with open(score_path, 'r', encoding='utf-8') as file:
        res_lis = json.load(file)
    for obj in res_lis:
        dimension = obj.get("dimension")

        if "gpt_ans" in obj:
            score = obj["gpt_ans"]['分数']
            try:
                score = int(score)
            except:
                try:
                    pattern = re.compile(r'\d+')
                    result = pattern.findall(score)
                    score = int(result[-1])
                except:
                    score = None
                    pass

        if score is not None and score in list(range(0, 11)):
            task_scores.get(dimension).append(score)

    wb = openpyxl.Workbook()
    sheet = wb.active
    headers = list(task_scores.keys())
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)
    for i, key in enumerate(task_scores.keys(), start=1):
        if len(task_scores[key]) != 0:
            average_score = sum(task_scores[key]) / len(task_scores[key])
        else:
            average_score = 0
        sheet.cell(row=2, column=i, value=average_score)
    wb.save(score_path_xlsx)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="结果评估")
    parser.add_argument("--parallel_batch", default=20, help="")
    parser.add_argument("--seed", default=3, help="")
    parser.add_argument("--exp_name", default='evaluation', help="")
    parser.add_argument("--eval_model", default='gpt-4o', help="")
    parser.add_argument("--rule", default='rules_exp', help="")
    parser.add_argument("--rule_name", default='rules_8', help="")

    args = parser.parse_args()

    random.seed(args.seed)
    answers = generate_gpt_prompt(
        answer_path="../{}.json".format(args.exp_name),
        rule_name=args.rule_name
    )
    answers = answers[:]

    prompt_path = "/prompt_gpt4_{}_{}_{}.json".format(args.rule_name, args.eval_model, args.exp_name)
    os.makedirs(os.path.dirname(prompt_path), exist_ok=True)

    eval_path = "/score_gpt4_{}_{}_{}.xlsx".format(args.rule_name, args.eval_model, args.exp_name)
    os.makedirs(os.path.dirname(eval_path), exist_ok=True)


    def process_answer(answer, model_name, answer_index):
        while True:
            response_origin = chat_with_gpt(
                query=answer["prompt"],
                key=key,
                model=model,
                json_mode=False)

            if response_origin:
                try:
                    response = response_origin.replace("json", '').replace('```', '')
                    response = response.strip()
                    data = eval(response)
                    answer["gpt_ans"] = data
                    assert '分数' in data, "'分数' in data"
                    log.info("answer_index: {}, gpt_ans: {}".format(answer_index, response))
                    return answer
                except:
                    log.error("WRONG!!!! response_origin:{}".format(response_origin))
            else:
                print("wrong connect")



    log.debug("共{}个任务".format(len(answers)))

    with concurrent.futures.ThreadPoolExecutor(max_workers=int(args.parallel_batch)) as executor:
     
        future_to_answer = {executor.submit(process_answer, answer, args.eval_model, answer_index): answer for answer_index, answer in
                            enumerate(answers)}
        for future in concurrent.futures.as_completed(future_to_answer):
            answer = future_to_answer[future]
            try:
                answer = future.result()
            except Exception as exc:
                log.error(f'Answer {answer["prompt"]} generated an exception: {exc}')

    with open(prompt_path, 'w', encoding="utf-8") as f:
        json.dump(answers, f, indent=4, ensure_ascii=False)

    get_gpt_score(prompt_path, eval_path, rule_name=args.rule)